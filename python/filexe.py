import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import apartments
import sys

wb = Workbook()

ws = wb.active  # grabs active worksheet

alph = ["A", "B", "C", "D", "E", "F"]  # Column Letters

# Customize each Top Row cell
ws["A1"] = "Name"
ws["A1"].font = Font(bold=True)

ws["B1"] = "Address"
ws["B1"].font = Font(bold=True)

ws["C1"] = "Cost"
ws["C1"].font = Font(bold=True)

ws["D1"] = "Room"
ws["D1"].font = Font(bold=True)

ws["E1"] = "Phone Number"
ws["E1"].font = Font(bold=True)

ws["F1"] = "info"
ws["F1"].font = Font(bold=True)

# resizeing each cell
for cell in alph:
    ws.column_dimensions[cell].width = len(ws[f"{cell}1"].value) + 5

wb.save(f"_{sys.argv[2]}/apartment_list.xlsx")  # save file


apart_data = apartments.aparment_info

# checks if theres a good amount of data to a a excel
if len(apart_data) < 20:
    sys.exit()
elif len(apart_data) >= 20:
    for row in apart_data:
        # adds each
        ws.append(row)

    # save the file
    wb.save(f"_{sys.argv[2]}/apartment_list.xlsx")
  


    workbook = openpyxl.load_workbook(f"_{sys.argv[2]}/apartment_list.xlsx")


    # Get the first sheet
    sheet = workbook.worksheets[0]

    # Create a list to store the values
    length = []
    column_count = sheet.max_column
    row_count = sheet.max_row
    # Iterate over the rows in the sheet
    for i in range(0, column_count):
        names = []
        for row in sheet:
            # Get the value of the first cell
            # in the row (the "Name" cell)
            name = row[i].value
            # Add the value to the list '
            names.append(len(name))
        names.sort(reverse=True)
        for cell in alph:
            ws.column_dimensions[cell].width = names[0]

    wb.save(f"_{sys.argv[2]}/apartment_list.xlsx")

